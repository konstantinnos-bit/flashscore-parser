import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime

# =====================
# НАСТРОЙКИ
# =====================
EXCEL_FILE = "Гол_во_втором_тайме_с_LIVE_HT.xlsx"

SHEET_NAME = "LIVE_HT"

FLASH_LIVE_URL = "https://www.flashscore.com/"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8"
}

# =====================
# ЗАГРУЗКА EXCEL
# =====================
wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

# очищаем старые данные, НЕ трогая заголовки
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row)

# =====================
# ПАРСИНГ FLASHScore
# =====================
response = requests.get(FLASH_LIVE_URL, headers=HEADERS, timeout=20)
soup = BeautifulSoup(response.text, "html.parser")

rows_added = 0
match_number = 1

matches = soup.select("div.event__match")

for match in matches:
    try:
        # счёт первого тайма
        score = match.select_one(".event__scores")
        if not score:
            continue

        ht_score = score.text.strip()

        # фильтр HT
        if ht_score not in ["0:0", "1:0", "0:1"]:
            continue

        team1 = match.select_one(".event__participant--home").text.strip()
        team2 = match.select_one(".event__participant--away").text.strip()

        # ---- ПОКА FlashScore не даёт всё в HTML ----
        # коэффициенты / владение / удары
        # ставим None — формулы Excel это переварят
        p1 = None
        p2 = None
        tm45 = None
        poss1 = None
        poss2 = None
        shots1 = None
        shots2 = None

        # последние 2 игры (будет отдельный этап)
        k1_g1_z = k1_g1_p = k1_g2_z = k1_g2_p = None
        k2_g1_z = k2_g1_p = k2_g2_z = k2_g2_p = None

        ws.append([
            match_number,
            team1,
            team2,
            ht_score.replace(":", "-"),
            p1,
            p2,
            tm45,
            poss1,
            poss2,
            shots1,
            shots2,
            k1_g1_z, k1_g1_p, k1_g2_z, k1_g2_p,
            k2_g1_z, k2_g1_p, k2_g2_z, k2_g2_p
        ])

        match_number += 1
        rows_added += 1

    except Exception:
        continue

# =====================
# СЛУЖЕБНАЯ ИНФОРМАЦИЯ
# =====================
ws["T1"] = "Обновлено:"
ws["T2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
ws["T3"] = f"Матчей найдено: {rows_added}"

# =====================
# СОХРАНЕНИЕ
# =====================
wb.save(EXCEL_FILE)
